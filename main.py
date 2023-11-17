import openpyxl as op

filename = 'Марки и модели для Инкар.xlsx'

filename2 = 'car_models.xlsx'

wb = op.load_workbook(filename, data_only=True)

wb2 = op.load_workbook(filename2, data_only=True)

sheet = wb.active

sheet2 = wb2.active

max_rows = sheet.max_row

max_rows2 = sheet2.max_row

list = []

for i in range(2, max_rows):
    c = 0
    first = sheet.cell(row=i, column=2).value
    for j in range(2, max_rows2):
        second = sheet2.cell(row=j, column=5).value
        if first == second:
            c += 1
    if c == 0:
        r = [sheet.cell(row=i, column=1).value, first]
        list.append(r)

res = []

for i in range(len(list)):
    r = list[i]
    if r not in res:
        res.append(r)

filename3 = 'res.xlsx'

wb3 = op.load_workbook(filename3)

sheet3 = wb3.active

for i in range(len(res)):
    a = 'A' + str(i+2)
    sheet3[a] = res[i][0]
    b = 'B' + str(i+2)
    sheet3[b] = res[i][1]
    # a = 'A' + str(i + 2)
    # sheet3[a] = ' '
    # b = 'B' + str(i + 2)
    # sheet3[b] = ' '
    # c = 'C' + str(i + 2)
    # sheet3[c] = ' '

wb3.save(filename3)